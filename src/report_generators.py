"""
Report Generators

This module handles the generation of various report formats (CSV, Excel, HTML, JSON)
from the collected IAM Identity Center data.
"""

import csv
import json
from datetime import datetime
from typing import List, Set, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .data_models import (
    ACCOUNT_ANALYSIS_CSV_FIELDNAMES,
    ANALYSIS_CSV_BASE_FIELDNAMES,
    CSV_FIELDNAMES,
    RISK_ANALYSIS_CSV_FIELDNAMES,
    AccessLevel,
    AccountAnalysis,
    RiskAnalysis,
    RiskLevel,
    UserAccountRoleGroup,
    UserAnalysis,
    UserSummary,
)


class ReportGenerator:
    """Generates reports in multiple formats."""

    def __init__(self, output_prefix: str = "iam_identity_center_report"):
        self.output_prefix = output_prefix
        self.start_time = datetime.now()

    def generate_all_reports(
        self,
        user_account_roles: List[UserAccountRoleGroup],
        user_summaries: List[UserSummary],
    ):
        """Generate all report formats."""
        print("Generating reports...")

        self.generate_csv_report(user_account_roles)
        self.generate_excel_report(user_account_roles)
        self.generate_html_report(user_account_roles)
        self.generate_json_report(user_summaries)
        self.generate_analysis_csv_report(user_account_roles)
        self.generate_accounts_csv_report(user_account_roles)
        self.generate_risk_analysis_csv_report(user_account_roles)

        self._print_completion_summary()

    def generate_csv_report(self, user_account_roles: List[UserAccountRoleGroup]):
        """Generate CSV report."""
        filename = f"{self.output_prefix}.csv"

        with open(filename, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=CSV_FIELDNAMES)
            writer.writeheader()

            for uar in user_account_roles:
                row = uar.to_dict()
                # Filter only CSV fields
                csv_row = {k: v for k, v in row.items() if k in CSV_FIELDNAMES}
                writer.writerow(csv_row)

        print(f"CSV file {filename} generated.")

    def generate_excel_report(self, user_account_roles: List[UserAccountRoleGroup]):
        """Generate Excel report with formatting and multiple analysis tabs."""
        filename = f"{self.output_prefix}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"  # Renamed from "IAM Identity Center" to "Data"

        # Add headers
        ws.append(CSV_FIELDNAMES)

        # Add data rows
        for uar in user_account_roles:
            row_data = uar.to_dict()
            ws.append([row_data.get(col, "") for col in CSV_FIELDNAMES])

        # Apply formatting
        self._format_excel_worksheet(ws)

        # Add analysis tabs
        self._add_users_analysis_worksheet(
            wb, user_account_roles
        )  # Renamed from "Analyse" to "Users"
        self._add_accounts_analysis_worksheet(wb, user_account_roles)
        self._add_risk_analysis_worksheet(wb, user_account_roles)
        self._add_summary_worksheet(wb, user_account_roles)

        wb.save(filename)
        print(f"XLSX file {filename} generated with multiple analysis tabs.")

    def generate_html_report(self, user_account_roles: List[UserAccountRoleGroup]):
        """Generate interactive multi-tab HTML report mirroring Excel structure."""
        filename = f"{self.output_prefix}.html"

        # Prepare all analysis data
        unique_users, classifications = self._extract_unique_users_with_access(
            user_account_roles
        )
        account_analyses = self._extract_account_analyses(user_account_roles)
        risk_analyses = self._extract_risk_analyses(user_account_roles)

        # Generate HTML content for each tab
        data_html = self._generate_data_table_html(user_account_roles)
        users_html = self._generate_users_table_html(unique_users, classifications)
        accounts_html = self._generate_accounts_table_html(account_analyses)
        risk_html = self._generate_risk_table_html(risk_analyses)
        summary_html = self._generate_summary_html(user_account_roles)

        # Generate complete multi-tab HTML
        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>IAM Identity Center Report</title>
    <link rel="stylesheet" href="https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css">
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 0;
            background-color: #f5f5f5;
        }}
        .container {{
            background-color: white;
            margin: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        h1 {{
            color: #333;
            text-align: center;
            margin: 20px 0;
            padding: 0 20px;
        }}

        /* Tab Styles */
        .tab-container {{
            border-bottom: 1px solid #ddd;
        }}
        .tab-buttons {{
            display: flex;
            background-color: #f8f9fa;
            margin: 0;
            padding: 0;
            list-style: none;
        }}
        .tab-button {{
            background-color: #f8f9fa;
            border: none;
            padding: 12px 24px;
            cursor: pointer;
            border-bottom: 3px solid transparent;
            font-size: 14px;
            font-weight: 500;
            transition: all 0.3s ease;
        }}
        .tab-button:hover {{
            background-color: #e9ecef;
        }}
        .tab-button.active {{
            background-color: white;
            border-bottom-color: #4CAF50;
            color: #4CAF50;
        }}

        .tab-content {{
            display: none;
            padding: 20px;
        }}
        .tab-content.active {{
            display: block;
        }}

        /* Table Styles */
        .table-container {{
            overflow-x: auto;
            margin-bottom: 20px;
        }}
        table.dataTable {{
            width: 100% !important;
            margin: 0 auto;
        }}
        table.dataTable th {{
            background-color: #4CAF50;
            color: white;
            font-weight: bold;
            text-align: center;
        }}
        table.dataTable td {{
            text-align: center;
            vertical-align: middle;
        }}
        .dataTables_wrapper .dataTables_length,
        .dataTables_wrapper .dataTables_filter,
        .dataTables_wrapper .dataTables_info,
        .dataTables_wrapper .dataTables_paginate {{
            margin: 10px 0;
        }}

        /* Summary Styles */
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
            margin-bottom: 20px;
        }}
        .summary-card {{
            background-color: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
            border-left: 4px solid #4CAF50;
        }}
        .summary-card h3 {{
            margin: 0 0 15px 0;
            color: #333;
        }}
        .summary-item {{
            display: flex;
            justify-content: space-between;
            margin: 8px 0;
            padding: 4px 0;
            border-bottom: 1px solid #e9ecef;
        }}
        .summary-item:last-child {{
            border-bottom: none;
        }}
        .summary-label {{
            font-weight: 500;
        }}
        .summary-value {{
            color: #4CAF50;
            font-weight: bold;
        }}

        .footer {{
            margin-top: 30px;
            text-align: center;
            color: #666;
            font-size: 0.9em;
            padding: 20px;
            background-color: #f8f9fa;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>IAM Identity Center Report</h1>

        <div class="tab-container">
            <div class="tab-buttons">
                <button class="tab-button active" onclick="showTab('data')">Data</button>
                <button class="tab-button" onclick="showTab('users')">Users</button>
                <button class="tab-button" onclick="showTab('accounts')">Accounts</button>
                <button class="tab-button" onclick="showTab('risk')">Risk Analysis</button>
                <button class="tab-button" onclick="showTab('summary')">Summary</button>
            </div>
        </div>

        <div id="data" class="tab-content active">
            <h2>Detailed Data</h2>
            <div class="table-container">
                {data_html}
            </div>
        </div>

        <div id="users" class="tab-content">
            <h2>Users Analysis</h2>
            <div class="table-container">
                {users_html}
            </div>
        </div>

        <div id="accounts" class="tab-content">
            <h2>Accounts Analysis</h2>
            <div class="table-container">
                {accounts_html}
            </div>
        </div>

        <div id="risk" class="tab-content">
            <h2>Risk Analysis</h2>
            <div class="table-container">
                {risk_html}
            </div>
        </div>

        <div id="summary" class="tab-content">
            <h2>Summary</h2>
            {summary_html}
        </div>

        <div class="footer">
            <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>

    <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
    <script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
    <script>
        function showTab(tabName) {{
            // Hide all tab contents
            const tabContents = document.querySelectorAll('.tab-content');
            tabContents.forEach(content => content.classList.remove('active'));

            // Remove active class from all buttons
            const tabButtons = document.querySelectorAll('.tab-button');
            tabButtons.forEach(button => button.classList.remove('active'));

            // Show selected tab content
            document.getElementById(tabName).classList.add('active');

            // Add active class to clicked button
            event.target.classList.add('active');

            // Reinitialize DataTables for the active tab
            setTimeout(() => {{
                const activeTable = document.querySelector(`#${{tabName}} table.display`);
                if (activeTable && !$.fn.DataTable.isDataTable(activeTable)) {{
                    $(activeTable).DataTable({{
                        "pageLength": 25,
                        "order": [[ 0, "asc" ]],
                        "responsive": true,
                        "scrollX": true
                    }});
                }}
            }}, 100);
        }}

        $(document).ready(function() {{
            // Initialize DataTable for the first tab
            $('#data table.display').DataTable({{
                "pageLength": 25,
                "order": [[ 0, "asc" ]],
                "responsive": true,
                "scrollX": true
            }});
        }});
    </script>
</body>
</html>
"""

        with open(filename, "w", encoding="utf-8") as f:
            f.write(html_content)

        print(f"Multi-tab HTML file {filename} generated.")

    def _generate_data_table_html(
        self, user_account_roles: List[UserAccountRoleGroup]
    ) -> str:
        """Generate HTML table for the main data tab."""
        # Prepare data for HTML
        html_rows = []
        for uar in user_account_roles:
            row_data = uar.to_dict()
            html_row = {}

            for col in CSV_FIELDNAMES:
                val = row_data.get(col, "")
                if isinstance(val, str):
                    html_row[col] = val.replace("\n", "<br>")
                else:
                    html_row[col] = val

            html_rows.append(html_row)

        # Create DataFrame and generate HTML table
        df = pd.DataFrame(html_rows, columns=CSV_FIELDNAMES)
        return df.to_html(index=False, escape=False, classes="display nowrap", border=0)

    def _generate_users_table_html(
        self, unique_users: List[UserAnalysis], classifications: Set[str]
    ) -> str:
        """Generate HTML table for the users analysis tab."""
        fieldnames = ANALYSIS_CSV_BASE_FIELDNAMES + sorted(classifications)

        html_rows = []
        for user_analysis in unique_users:
            row_data = user_analysis.to_dict(classifications=sorted(classifications))
            html_rows.append(row_data)

        df = pd.DataFrame(html_rows, columns=fieldnames)
        return df.to_html(index=False, escape=False, classes="display nowrap", border=0)

    def _generate_accounts_table_html(
        self, account_analyses: List[AccountAnalysis]
    ) -> str:
        """Generate HTML table for the accounts analysis tab."""
        html_rows = []
        for account_analysis in account_analyses:
            row_data = account_analysis.to_dict()
            # Convert email lists to HTML-friendly format
            for key in ["Admin Emails", "Read Write Emails", "Read Only Emails"]:
                if row_data[key]:
                    # Replace semicolons with line breaks for better HTML display
                    row_data[key] = row_data[key].replace("; ", "<br>")
            html_rows.append(row_data)

        df = pd.DataFrame(html_rows, columns=ACCOUNT_ANALYSIS_CSV_FIELDNAMES)
        return df.to_html(index=False, escape=False, classes="display nowrap", border=0)

    def _generate_risk_table_html(self, risk_analyses: List[RiskAnalysis]) -> str:
        """Generate HTML table for the risk analysis tab."""
        html_rows = []
        for risk_analysis in risk_analyses:
            html_rows.append(risk_analysis.to_dict())

        df = pd.DataFrame(html_rows, columns=RISK_ANALYSIS_CSV_FIELDNAMES)
        return df.to_html(index=False, escape=False, classes="display nowrap", border=0)

    def _generate_summary_html(
        self, user_account_roles: List[UserAccountRoleGroup]
    ) -> str:
        """Generate HTML content for the summary tab."""
        # Calculate summary metrics
        total_users = len(set(uar.user.username for uar in user_account_roles))
        total_accounts = len(set(uar.account.id for uar in user_account_roles))
        total_roles = len(set(uar.role.name for uar in user_account_roles))

        # Count by access level (active users only)
        admin_count = len(
            [
                uar
                for uar in user_account_roles
                if uar.role.access_level == AccessLevel.FULL_ADMIN
                and uar.user.status.lower() != "disabled"
            ]
        )
        read_write_count = len(
            [
                uar
                for uar in user_account_roles
                if uar.role.access_level == AccessLevel.READ_WRITE
                and uar.user.status.lower() != "disabled"
            ]
        )
        read_only_count = len(
            [
                uar
                for uar in user_account_roles
                if uar.role.access_level == AccessLevel.READ_ONLY
                and uar.user.status.lower() != "disabled"
            ]
        )

        # Count by risk level (active users only)
        high_risk_count = len(
            [
                uar
                for uar in user_account_roles
                if uar.role.risk_level in [RiskLevel.HIGH, RiskLevel.CRITICAL]
                and uar.user.status.lower() != "disabled"
            ]
        )

        # Count by classification
        classifications = {}
        for uar in user_account_roles:
            classification = uar.account.classification
            if classification not in classifications:
                classifications[classification] = set()
            classifications[classification].add(uar.user.username)

        # Generate summary HTML
        summary_html = f"""
        <div class="summary-grid">
            <div class="summary-card">
                <h3>Overall Statistics</h3>
                <div class="summary-item">
                    <span class="summary-label">Total Users:</span>
                    <span class="summary-value">{total_users}</span>
                </div>
                <div class="summary-item">
                    <span class="summary-label">Total Accounts:</span>
                    <span class="summary-value">{total_accounts}</span>
                </div>
                <div class="summary-item">
                    <span class="summary-label">Total Roles:</span>
                    <span class="summary-value">{total_roles}</span>
                </div>
            </div>

            <div class="summary-card">
                <h3>Access Level Distribution</h3>
                <div class="summary-item">
                    <span class="summary-label">Admin Users:</span>
                    <span class="summary-value">{admin_count}</span>
                </div>
                <div class="summary-item">
                    <span class="summary-label">Read Write Users:</span>
                    <span class="summary-value">{read_write_count}</span>
                </div>
                <div class="summary-item">
                    <span class="summary-label">Read Only Users:</span>
                    <span class="summary-value">{read_only_count}</span>
                </div>
            </div>

            <div class="summary-card">
                <h3>Risk Analysis</h3>
                <div class="summary-item">
                    <span class="summary-label">High/Critical Risk Users:</span>
                    <span class="summary-value">{high_risk_count}</span>
                </div>
            </div>

            <div class="summary-card">
                <h3>Users by Classification</h3>
        """

        for classification, users in sorted(classifications.items()):
            summary_html += f"""
                <div class="summary-item">
                    <span class="summary-label">{classification}:</span>
                    <span class="summary-value">{len(users)}</span>
                </div>
            """

        summary_html += """
            </div>
        </div>
        """

        return summary_html

    def generate_json_report(self, user_summaries: List[UserSummary]):
        """Generate JSON report."""
        filename = f"{self.output_prefix}.json"

        json_data = [summary.to_dict() for summary in user_summaries]

        with open(filename, "w", encoding="utf-8") as jf:
            json.dump(json_data, jf, ensure_ascii=False, indent=2)

        print(f"JSON file {filename} generated.")

    def _format_excel_worksheet(self, ws):
        """Apply formatting to Excel worksheet."""
        # Header formatting
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        # Auto-width columns and wrap text
        for col_idx, col in enumerate(CSV_FIELDNAMES, 1):
            max_length = len(col)

            for row_cells in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row_cells:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length

                    # Enable text wrapping
                    try:
                        cell.alignment = cell.alignment.copy(wrapText=True)
                    except AttributeError:
                        from openpyxl.styles import Alignment

                        cell.alignment = Alignment(wrapText=True)

            # Set column width (max 50 characters)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Freeze panes (first row and first column)
        ws.freeze_panes = ws.cell(row=2, column=2)

    def _create_html_template(self, html_table: str) -> str:
        """Create complete HTML template with DataTables."""
        return f"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>AWS IAM Identity Center Report</title>
<link rel="stylesheet" href="https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css">
<style>
table.dataTable thead th {{
    background-color: #4F81BD;
    color: white;
    font-weight: bold;
}}
table.dataTable tbody tr:nth-child(even) {{
    background-color: #f9f9f9;
}}
table.dataTable tbody tr:hover {{
    background-color: #e6f3ff;
}}
.container {{
    margin: 20px;
}}
h1 {{
    color: #4F81BD;
    font-family: Arial, sans-serif;
}}
</style>
</head>
<body>
<div class="container">
<h1>AWS IAM Identity Center Report</h1>
<p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
{html_table}
</div>

<script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
<script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
<script>
$(document).ready(function() {{
    $('.display').DataTable({{
        "pageLength": 25,
        "lengthMenu": [10, 25, 50, 100, -1],
        "order": [[ 0, "asc" ]],
        "columnDefs": [
            {{ "width": "15%", "targets": 0 }},
            {{ "width": "15%", "targets": 1 }},
            {{ "width": "20%", "targets": 2 }},
            {{ "width": "10%", "targets": 3 }},
            {{ "width": "15%", "targets": 4 }},
            {{ "width": "10%", "targets": 5 }},
            {{ "width": "5%", "targets": 6 }},
            {{ "width": "5%", "targets": 7 }},
            {{ "width": "5%", "targets": 8 }},
            {{ "width": "10%", "targets": 9 }}
        ]
    }});
}});
</script>
</body>
</html>
"""

    def _print_completion_summary(self):
        """Print completion summary with timing."""
        end_time = datetime.now()
        duration = (end_time - self.start_time).total_seconds()

        print(f"Script started at: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Script ended at:   {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Total duration:    {duration:.2f} seconds")

    def generate_analysis_csv_report(
        self, user_account_roles: List[UserAccountRoleGroup]
    ):
        """Generate CSV report with unique users and their access levels per account type."""
        filename = f"{self.output_prefix}_analysis.csv"

        # Extract unique users with access levels per classification
        unique_users, classifications = self._extract_unique_users_with_access(
            user_account_roles
        )

        # Build dynamic fieldnames
        fieldnames = ANALYSIS_CSV_BASE_FIELDNAMES + sorted(classifications)

        with open(filename, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            for user_analysis in unique_users:
                writer.writerow(
                    user_analysis.to_dict(classifications=sorted(classifications))
                )

        print(
            f"Analysis CSV file {filename} generated with {len(unique_users)} unique users across {len(classifications)} account types."
        )

    def _add_users_analysis_worksheet(
        self, workbook: Workbook, user_account_roles: List[UserAccountRoleGroup]
    ):
        """Add the 'Users' worksheet with unique users and access levels per account type."""
        # Create new worksheet
        ws_analysis = workbook.create_sheet(title="Users")

        # Extract unique users with access levels per classification
        unique_users, classifications = self._extract_unique_users_with_access(
            user_account_roles
        )

        # Build dynamic fieldnames
        fieldnames = ANALYSIS_CSV_BASE_FIELDNAMES + sorted(classifications)

        # Add headers
        ws_analysis.append(fieldnames)

        # Add data rows
        for user_analysis in unique_users:
            row_data = user_analysis.to_dict(classifications=sorted(classifications))
            ws_analysis.append([row_data.get(col, "") for col in fieldnames])

        # Apply formatting
        self._format_excel_worksheet(ws_analysis)

        print(
            f"Added 'Users' worksheet with {len(unique_users)} unique users across {len(classifications)} account types."
        )

    def generate_accounts_csv_report(
        self, user_account_roles: List[UserAccountRoleGroup]
    ):
        """Generate CSV report with account analysis."""
        filename = f"{self.output_prefix}_accounts.csv"

        # Extract account analysis
        account_analyses = self._extract_account_analyses(user_account_roles)

        with open(filename, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=ACCOUNT_ANALYSIS_CSV_FIELDNAMES)
            writer.writeheader()

            for account_analysis in account_analyses:
                writer.writerow(account_analysis.to_dict())

        print(
            f"Accounts CSV file {filename} generated with {len(account_analyses)} accounts."
        )

    def generate_risk_analysis_csv_report(
        self, user_account_roles: List[UserAccountRoleGroup]
    ):
        """Generate CSV report with risk analysis by classification."""
        filename = f"{self.output_prefix}_risk_analysis.csv"

        # Extract risk analysis
        risk_analyses = self._extract_risk_analyses(user_account_roles)

        with open(filename, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=RISK_ANALYSIS_CSV_FIELDNAMES)
            writer.writeheader()

            for risk_analysis in risk_analyses:
                writer.writerow(risk_analysis.to_dict())

        print(
            f"Risk analysis CSV file {filename} generated with {len(risk_analyses)} classifications."
        )

    def _add_accounts_analysis_worksheet(
        self, workbook: Workbook, user_account_roles: List[UserAccountRoleGroup]
    ):
        """Add the 'Accounts' worksheet with account analysis."""
        # Create new worksheet
        ws_accounts = workbook.create_sheet(title="Accounts")

        # Extract account analyses
        account_analyses = self._extract_account_analyses(user_account_roles)

        # Add headers
        ws_accounts.append(ACCOUNT_ANALYSIS_CSV_FIELDNAMES)

        # Add data rows
        for account_analysis in account_analyses:
            row_data = account_analysis.to_dict()
            ws_accounts.append(
                [row_data.get(col, "") for col in ACCOUNT_ANALYSIS_CSV_FIELDNAMES]
            )

        # Apply formatting
        self._format_excel_worksheet(ws_accounts)

        print(f"Added 'Accounts' worksheet with {len(account_analyses)} accounts.")

    def _add_risk_analysis_worksheet(
        self, workbook: Workbook, user_account_roles: List[UserAccountRoleGroup]
    ):
        """Add the 'Risk Analysis' worksheet."""
        # Create new worksheet
        ws_risk = workbook.create_sheet(title="Risk Analysis")

        # Extract risk analyses
        risk_analyses = self._extract_risk_analyses(user_account_roles)

        # Add headers
        ws_risk.append(RISK_ANALYSIS_CSV_FIELDNAMES)

        # Add data rows
        for risk_analysis in risk_analyses:
            row_data = risk_analysis.to_dict()
            ws_risk.append(
                [row_data.get(col, "") for col in RISK_ANALYSIS_CSV_FIELDNAMES]
            )

        # Apply formatting
        self._format_excel_worksheet(ws_risk)

        print(
            f"Added 'Risk Analysis' worksheet with {len(risk_analyses)} classifications."
        )

    def _add_summary_worksheet(
        self, workbook: Workbook, user_account_roles: List[UserAccountRoleGroup]
    ):
        """Add a 'Summary' worksheet with key metrics."""
        # Create new worksheet
        ws_summary = workbook.create_sheet(title="Summary")

        # Calculate summary metrics
        total_users = len(set(uar.user.username for uar in user_account_roles))
        total_accounts = len(set(uar.account.id for uar in user_account_roles))
        total_roles = len(set(uar.role.name for uar in user_account_roles))

        # Count by access level
        admin_count = len(
            [
                uar
                for uar in user_account_roles
                if uar.role.access_level == AccessLevel.FULL_ADMIN
                and uar.user.status.lower() != "disabled"
            ]
        )
        read_write_count = len(
            [
                uar
                for uar in user_account_roles
                if uar.role.access_level == AccessLevel.READ_WRITE
                and uar.user.status.lower() != "disabled"
            ]
        )
        read_only_count = len(
            [
                uar
                for uar in user_account_roles
                if uar.role.access_level == AccessLevel.READ_ONLY
                and uar.user.status.lower() != "disabled"
            ]
        )

        # Count by risk level
        high_risk_count = len(
            [
                uar
                for uar in user_account_roles
                if uar.role.risk_level in [RiskLevel.HIGH, RiskLevel.CRITICAL]
                and uar.user.status.lower() != "disabled"
            ]
        )

        # Count by classification
        classifications = {}
        for uar in user_account_roles:
            classification = uar.account.classification
            if classification not in classifications:
                classifications[classification] = set()
            classifications[classification].add(uar.user.username)

        # Add summary data
        summary_data = [
            ["Metric", "Value"],
            ["Total Users", total_users],
            ["Total Accounts", total_accounts],
            ["Total Roles", total_roles],
            ["", ""],
            ["Access Level Distribution", ""],
            ["Admin Users", admin_count],
            ["Read Write Users", read_write_count],
            ["Read Only Users", read_only_count],
            ["", ""],
            ["Risk Analysis", ""],
            ["High/Critical Risk Users", high_risk_count],
            ["", ""],
            ["Users by Classification", ""],
        ]

        for classification, users in sorted(classifications.items()):
            summary_data.append([f"{classification} Users", len(users)])

        # Add all rows
        for row in summary_data:
            ws_summary.append(row)

        # Apply formatting
        self._format_excel_worksheet(ws_summary)

        print("Added 'Summary' worksheet with key metrics.")

    def _extract_account_analyses(
        self, user_account_roles: List[UserAccountRoleGroup]
    ) -> List[AccountAnalysis]:
        """Extract account analyses from user account roles."""
        account_data = {}

        for uar in user_account_roles:
            account_key = uar.account.id

            if account_key not in account_data:
                account_data[account_key] = {
                    "account": uar.account,
                    "users": set(),
                    "admin_emails": set(),
                    "read_write_emails": set(),
                    "read_only_emails": set(),
                }

            data = account_data[account_key]
            data["users"].add(uar.user.username)

            # Determine access level considering disabled users
            if uar.user.status.lower() == "disabled":
                access_level = AccessLevel.NO_ACCESS
            else:
                access_level = uar.role.access_level

            # Add email to appropriate access level list
            if access_level == AccessLevel.FULL_ADMIN:
                data["admin_emails"].add(uar.user.email)
            elif access_level == AccessLevel.READ_WRITE:
                data["read_write_emails"].add(uar.user.email)
            elif access_level == AccessLevel.READ_ONLY:
                data["read_only_emails"].add(uar.user.email)

        # Convert to AccountAnalysis objects
        analyses = []
        for account_key, data in account_data.items():
            analysis = AccountAnalysis(
                account_name=data["account"].name,
                account_id=data["account"].id,
                classification=data["account"].classification,
                user_count=len(data["users"]),
                admin_emails=list(data["admin_emails"]),
                read_write_emails=list(data["read_write_emails"]),
                read_only_emails=list(data["read_only_emails"]),
            )
            analyses.append(analysis)

        # Sort by account name
        return sorted(analyses, key=lambda x: x.account_name)

    def _extract_risk_analyses(
        self, user_account_roles: List[UserAccountRoleGroup]
    ) -> List[RiskAnalysis]:
        """Extract risk analyses by classification."""
        classification_data = {}

        for uar in user_account_roles:
            classification = uar.account.classification

            if classification not in classification_data:
                classification_data[classification] = {
                    "accounts": set(),
                    "users": set(),
                    "admin_users": set(),
                    "high_risk_users": set(),
                    "critical_risk_users": set(),
                }

            data = classification_data[classification]
            data["accounts"].add(uar.account.id)

            # Only count active users
            if uar.user.status.lower() != "disabled":
                data["users"].add(uar.user.username)

                if uar.role.access_level == AccessLevel.FULL_ADMIN:
                    data["admin_users"].add(uar.user.username)

                if uar.role.risk_level == RiskLevel.HIGH:
                    data["high_risk_users"].add(uar.user.username)
                elif uar.role.risk_level == RiskLevel.CRITICAL:
                    data["critical_risk_users"].add(uar.user.username)

        # Convert to RiskAnalysis objects
        analyses = []
        for classification, data in classification_data.items():
            analysis = RiskAnalysis(
                classification=classification,
                total_accounts=len(data["accounts"]),
                total_users=len(data["users"]),
                admin_users=len(data["admin_users"]),
                high_risk_users=len(data["high_risk_users"]),
                critical_risk_users=len(data["critical_risk_users"]),
            )
            analyses.append(analysis)

        # Sort by classification name
        return sorted(analyses, key=lambda x: x.classification)

    def _extract_unique_users_with_access(
        self, user_account_roles: List[UserAccountRoleGroup]
    ) -> Tuple[List[UserAnalysis], Set[str]]:
        """Extract unique users with their highest access level per account classification."""
        unique_users_dict = {}
        all_classifications = set()

        # Access level priority (higher number = higher access)
        access_priority = {
            AccessLevel.NO_ACCESS.value: 0,
            AccessLevel.READ_ONLY.value: 1,
            AccessLevel.READ_WRITE.value: 2,
            AccessLevel.FULL_ADMIN.value: 3,
        }

        for uar in user_account_roles:
            user_key = uar.user.username
            classification = uar.account.classification
            all_classifications.add(classification)

            # Get current access level, considering disabled users
            current_access = (
                AccessLevel.NO_ACCESS.value
                if uar.user.status.lower() == "disabled"
                else uar.role.access_level.value
            )

            # Initialize user if not exists
            if user_key not in unique_users_dict:
                unique_users_dict[user_key] = UserAnalysis(
                    username=uar.user.username,
                    email=uar.user.email,
                    access_by_classification={},
                )

            user_analysis = unique_users_dict[user_key]

            # Update access level for this classification if it's higher than current
            existing_access = user_analysis.access_by_classification.get(
                classification, AccessLevel.NO_ACCESS.value
            )
            current_priority = access_priority.get(current_access, 0)
            existing_priority = access_priority.get(existing_access, 0)

            if current_priority > existing_priority:
                user_analysis.access_by_classification[classification] = current_access

        # Sort by username for consistent output
        sorted_users = sorted(unique_users_dict.values(), key=lambda x: x.username)
        return sorted_users, all_classifications

    def _extract_unique_users(
        self, user_account_roles: List[UserAccountRoleGroup]
    ) -> List[UserAnalysis]:
        """Extract unique users from user account roles (legacy method for backward compatibility)."""
        users, _ = self._extract_unique_users_with_access(user_account_roles)
        return users
