#!/usr/bin/env python3
"""
AWS IAM Identity Center (SSO) Reporting Script
Copyright (C) 2025 Cyril Feraudet, Nuagic

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program. If not, see <https://www.gnu.org/licenses/>.

See README.md for full usage and documentation.
"""
__author__ = "Cyril Feraudet"
import boto3
import csv
from datetime import datetime
from openpyxl import Workbook

# Set the AWS profile to use
# Utilise la configuration d'authentification par défaut de boto3
# (variables d'env, AWS_PROFILE, SSO, etc)
session = boto3.Session()

# Required clients
esso_admin = session.client("sso-admin")
identitystore = session.client("identitystore")

start_time = datetime.now()
print(f"Script started at {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
print("Fetching IAM Identity Center instance...")
# Get the Identity Center (SSO) instance
response = esso_admin.list_instances()
if not response["Instances"]:
    raise Exception("No IAM Identity Center instance found.")
instance = response["Instances"][0]
instance_arn = instance["InstanceArn"]
identity_store_id = instance["IdentityStoreId"]

# --- Caching Section ---

# Cache all users
print("Listing all users...")
users = []
paginator = identitystore.get_paginator("list_users")
for page in paginator.paginate(IdentityStoreId=identity_store_id):
    users.extend(page["Users"])
print(f"Found {len(users)} users.")

# Cache all groups and group memberships
print("Listing all groups and memberships...")
groups = []
group_id_to_name = {}
group_id_to_members = {}
paginator = identitystore.get_paginator("list_groups")
for page in paginator.paginate(IdentityStoreId=identity_store_id):
    groups.extend(page["Groups"])
for group in groups:
    group_id = group["GroupId"]
    group_id_to_name[group_id] = group["DisplayName"]
    group_id_to_members[group_id] = set()
    # Cache all members of the group
    paginator = identitystore.get_paginator("list_group_memberships")
    for page in paginator.paginate(IdentityStoreId=identity_store_id, GroupId=group_id):
        for membership in page["GroupMemberships"]:
            member = membership["MemberId"]
            if "UserId" in member:
                group_id_to_members[group_id].add(member["UserId"])
print(f"Found {len(groups)} groups.")

# Cache all AWS accounts
print("Listing all AWS accounts...")
org_client = session.client("organizations")
org_paginator = org_client.get_paginator("list_accounts")
all_accounts = []
account_id_to_name = {}
for page in org_paginator.paginate():
    all_accounts.extend(page["Accounts"])
for account in all_accounts:
    account_id_to_name[account["Id"]] = account["Name"]
print(f"Found {len(all_accounts)} accounts.")

# Cache all PermissionSets
print("Listing all PermissionSets...")
ps_paginator = esso_admin.get_paginator("list_permission_sets")
permission_sets = []
permission_set_arn_to_name = {}
for page in ps_paginator.paginate(InstanceArn=instance_arn):
    permission_sets.extend(page["PermissionSets"])
for arn in permission_sets:
    resp = esso_admin.describe_permission_set(
        InstanceArn=instance_arn, PermissionSetArn=arn
    )
    permission_set_arn_to_name[arn] = resp["PermissionSet"]["Name"]
print(f"Found {len(permission_sets)} PermissionSets.")

# Cache permission analysis for each PermissionSet
print("Analyzing permissions for each PermissionSet...")
permission_set_arn_to_access_level = {}


def analyze_permission_set_access_level(permission_set_arn):
    """
    Analyze a Permission Set to determine its access level and scores:
    - 'read-only': Only read/list/describe actions
    - 'read-write': Read + write/modify actions, but not admin
    - 'full-admin': Administrative access or wildcard permissions
    
    Returns: (access_level, read_score, write_score, admin_score)
    Scores are 0-10 where 10 = maximum level for that category
    """
    try:
        # Get managed policies
        managed_policies = []
        paginator = esso_admin.get_paginator("list_managed_policies_in_permission_set")
        for page in paginator.paginate(
            InstanceArn=instance_arn, PermissionSetArn=permission_set_arn
        ):
            managed_policies.extend(page["AttachedManagedPolicies"])

        # Get inline policy
        inline_policy = None
        try:
            resp = esso_admin.get_inline_policy_for_permission_set(
                InstanceArn=instance_arn, PermissionSetArn=permission_set_arn
            )
            inline_policy = resp.get("InlinePolicy")
        except esso_admin.exceptions.ResourceNotFoundException:
            pass

        # Initialize scores (0-10 scale)
        read_score = 0
        write_score = 0
        admin_score = 0
        
        has_write_actions = False
        has_admin_actions = False
        has_wildcard_actions = False

        # Check managed policies
        for policy in managed_policies:
            policy_arn = policy["Arn"]
            policy_name = policy["Name"]
            
            # Known admin policies (score 10)
            admin_policies = [
                "AdministratorAccess",
                "PowerUserAccess", 
                "IAMFullAccess",
                "OrganizationsFullAccess"
            ]
            
            if any(admin_policy in policy_name for admin_policy in admin_policies):
                has_admin_actions = True
                admin_score = 10
                write_score = 10
                read_score = 10
                break
                
            # Check for read-only policies (score 8-10 for read, 0 for write/admin)
            if "ReadOnly" in policy_name or "ViewOnly" in policy_name:
                read_score = max(read_score, 8)
                continue
            else:
                # Other policies likely have write permissions
                has_write_actions = True
                read_score = max(read_score, 6)
                write_score = max(write_score, 6)

        # Basic inline policy analysis (simplified)
        if inline_policy:
            import json
            try:
                policy_doc = json.loads(inline_policy)
                for statement in policy_doc.get("Statement", []):
                    if isinstance(statement, dict):
                        actions = statement.get("Action", [])
                        if isinstance(actions, str):
                            actions = [actions]
                        
                        for action in actions:
                            if action == "*" or action.endswith(":*"):
                                has_wildcard_actions = True
                                has_admin_actions = True
                                admin_score = 10
                                write_score = 10
                                read_score = 10
                                break
                            elif any(admin_action in action.lower() for admin_action in 
                                   ["create", "delete", "put", "update", "modify", "attach", "detach"]):
                                has_write_actions = True
                                write_score = max(write_score, 7)
                                read_score = max(read_score, 5)
                            elif any(admin_action in action.lower() for admin_action in 
                                   ["*", "admin", "full", "manage"]):
                                has_admin_actions = True
                                admin_score = max(admin_score, 8)
                                write_score = max(write_score, 8)
                                read_score = max(read_score, 8)
                                break
                            elif any(read_action in action.lower() for read_action in
                                   ["list", "describe", "get", "read"]):
                                read_score = max(read_score, 4)
            except (json.JSONDecodeError, KeyError):
                pass

        # Ensure minimum scores based on detected actions
        if has_admin_actions or has_wildcard_actions:
            admin_score = max(admin_score, 8)
            write_score = max(write_score, 8)
            read_score = max(read_score, 8)
        elif has_write_actions:
            write_score = max(write_score, 6)
            read_score = max(read_score, 5)
        elif read_score == 0:
            # Default minimum read score if no specific actions detected
            read_score = 3

        # Determine access level
        if has_admin_actions or has_wildcard_actions:
            access_level = "full-admin"
        elif has_write_actions:
            access_level = "read-write"
        else:
            access_level = "read-only"
            
        return access_level, read_score, write_score, admin_score

    except Exception as e:
        print(f"Warning: Could not analyze permissions for {permission_set_arn}: {e}")
        return "unknown", 0, 0, 0


# Analyze all permission sets
permission_set_arn_to_scores = {}
for arn in permission_sets:
    access_level, read_score, write_score, admin_score = analyze_permission_set_access_level(arn)
    permission_set_arn_to_access_level[arn] = access_level
    permission_set_arn_to_scores[arn] = {
        'read_score': read_score,
        'write_score': write_score, 
        'admin_score': admin_score
    }

print(f"Analyzed permissions for {len(permission_sets)} PermissionSets.")

# Cache all assignments for each (account, permission_set)
print("Caching all assignments for each account and PermissionSet...")
assignments = {}  # (account_id, permission_set_arn) -> list of assignments
for permission_set_arn in permission_sets:
    for account in all_accounts:
        account_id = account["Id"]
        key = (account_id, permission_set_arn)
        assignments[key] = []
        assign_paginator = esso_admin.get_paginator("list_account_assignments")
        for page in assign_paginator.paginate(
            InstanceArn=instance_arn,
            AccountId=account_id,
            PermissionSetArn=permission_set_arn,
        ):
            assignments[key].extend(page["AccountAssignments"])
print("Assignment cache built.")

# --- Helper functions using cache ---


def get_user_groups(user_id):
    user_groups = []
    for group_id, members in group_id_to_members.items():
        if user_id in members:
            user_groups.append(group_id_to_name[group_id])
    return user_groups


def get_user_account_roles(user_id):
    user_accounts = set()
    user_roles = set()
    user_groups = set()
    for group_id, members in group_id_to_members.items():
        if user_id in members:
            user_groups.add(group_id)
    for (account_id, permission_set_arn), assigns in assignments.items():
        for assignment in assigns:
            if (
                assignment["PrincipalType"] == "USER"
                and assignment["PrincipalId"] == user_id
            ):
                user_accounts.add(account_id)
                user_roles.add(permission_set_arn)
            if (
                assignment["PrincipalType"] == "GROUP"
                and assignment["PrincipalId"] in user_groups
            ):
                user_accounts.add(account_id)
                user_roles.add(permission_set_arn)
    return user_accounts, user_roles


def get_permission_set_name(permission_set_arn):
    return permission_set_arn_to_name.get(permission_set_arn, permission_set_arn)


def get_account_name(account_id):
    return account_id_to_name.get(account_id, account_id)


print("Generating CSV and XLSX report...")
fieldnames = ["User", "Groups", "AWS Accounts", "Role Details", "Max Read Score", "Max Write Score", "Max Admin Score", "Highest Risk Level"]
rows = []
for idx, user in enumerate(users, 1):
    user_id = user["UserId"]
    username = user.get("UserName", user.get("DisplayName", user_id))
    print(f"[{idx}/{len(users)}] Processing user: {username}")
    groups = get_user_groups(user_id)
    # Build accounts and roles per account for this user
    accounts_roles = {}
    # Pass 1: collect all assignments (direct and via groups)
    user_accounts = set()
    user_groups_ids = set()
    for group_id, members in group_id_to_members.items():
        if user_id in members:
            user_groups_ids.add(group_id)
    for (account_id, permission_set_arn), assigns in assignments.items():
        for assignment in assigns:
            if (
                assignment["PrincipalType"] == "USER"
                and assignment["PrincipalId"] == user_id
            ) or (
                assignment["PrincipalType"] == "GROUP"
                and assignment["PrincipalId"] in user_groups_ids
            ):
                if account_id not in accounts_roles:
                    accounts_roles[account_id] = set()
                accounts_roles[account_id].add(permission_set_arn)
    # Format AWS Accounts column: one line per account, with roles in
    # parentheses
    aws_accounts_lines = []
    for acc_id in sorted(accounts_roles.keys(), key=lambda x: get_account_name(x)):
        acc_name = get_account_name(acc_id)
        role_names = [
            get_permission_set_name(arn) for arn in sorted(accounts_roles[acc_id])
        ]
        aws_accounts_lines.append(f"{acc_name} ({', '.join(role_names)})")
    # Format Groups with line separator
    groups_str = "\n".join(groups)
    aws_accounts_str = "\n".join(aws_accounts_lines)
    if username == "cyril.feraudet@nuant.com":
        print("[DEBUG] --- Cyril account/role mapping ---")
        for acc_id, role_arns in accounts_roles.items():
            roles = [get_permission_set_name(arn) for arn in role_arns]
            print(
                f"[DEBUG] Cyril: AccountId={acc_id}, "
                f"AccountName={get_account_name(acc_id)}, "
                f"Roles={roles}"
            )
        print("[DEBUG] --- All known accounts ---")
        for acc_id in sorted(account_id_to_name.keys()):
            print(
                f"[DEBUG] Known: AccountId={acc_id}, "
                f"AccountName={account_id_to_name[acc_id]}"
            )
        print(f"[DEBUG] Cyril - Final CSV Accounts: " f"{aws_accounts_lines}")
    # Calculate user's maximum scores and risk level
    max_read_score = 0
    max_write_score = 0
    max_admin_score = 0
    role_details = []
    
    # Structure pour JSON : comptes avec rôles en listes et niveaux d'accès
    aws_accounts_json = []
    for acc_id in sorted(accounts_roles.keys(), key=lambda x: get_account_name(x)):
        acc_name = get_account_name(acc_id)
        roles_with_access_level = []
        for arn in sorted(accounts_roles[acc_id]):
            role_name = get_permission_set_name(arn)
            access_level = permission_set_arn_to_access_level.get(arn, "unknown")
            scores = permission_set_arn_to_scores.get(arn, {'read_score': 0, 'write_score': 0, 'admin_score': 0})
            
            # Update maximum scores
            max_read_score = max(max_read_score, scores['read_score'])
            max_write_score = max(max_write_score, scores['write_score'])
            max_admin_score = max(max_admin_score, scores['admin_score'])
            
            # Add detailed role information for CSV/XLSX
            role_details.append(f"{role_name} ({access_level}: R{scores['read_score']}/W{scores['write_score']}/A{scores['admin_score']})")
            
            roles_with_access_level.append({
                "name": role_name,
                "access_level": access_level,
                "read_score": scores['read_score'],
                "write_score": scores['write_score'],
                "admin_score": scores['admin_score']
            })
        aws_accounts_json.append({
            "account_name": acc_name, 
            "account_id": acc_id, 
            "roles": roles_with_access_level
        })

    # Determine highest risk level
    if max_admin_score >= 8:
        highest_risk = "CRITICAL"
    elif max_admin_score >= 5 or max_write_score >= 8:
        highest_risk = "HIGH"
    elif max_write_score >= 5:
        highest_risk = "MEDIUM"
    elif max_read_score >= 5:
        highest_risk = "LOW"
    else:
        highest_risk = "MINIMAL"

    role_details_str = "\n".join(role_details)

    row = {
        "User": username,
        "Groups": groups_str,
        "AWS Accounts": aws_accounts_str,
        "Role Details": role_details_str,
        "Max Read Score": max_read_score,
        "Max Write Score": max_write_score,
        "Max Admin Score": max_admin_score,
        "Highest Risk Level": highest_risk,
        "_groups_list": groups,  # pour JSON
        "_aws_accounts_json": aws_accounts_json  # pour JSON avec structure
    }
    rows.append(row)

# Write CSV
with open(
    "iam_identity_center_report.csv", "w", newline="", encoding="utf-8"
) as csvfile:
    csv_fieldnames = ["User", "Groups", "AWS Accounts", "Role Details", "Max Read Score", "Max Write Score", "Max Admin Score", "Highest Risk Level"]
    writer = csv.DictWriter(csvfile, fieldnames=csv_fieldnames)
    writer.writeheader()
    for row in rows:
        # Ne garder que les champs du CSV (exclure les champs internes
        # pour JSON)
        csv_row = {k: v for k, v in row.items() if k in csv_fieldnames}
        writer.writerow(csv_row)

# Write XLSX

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "IAM Identity Center"
xlsx_fieldnames = ["User", "Groups", "AWS Accounts", "Role Details", "Max Read Score", "Max Write Score", "Max Admin Score", "Highest Risk Level"]
ws.append(xlsx_fieldnames)
# Ajout des lignes de données après l'en-tête
for row in rows:
    xlsx_row = {k: v for k, v in row.items() if k in xlsx_fieldnames}
    ws.append([xlsx_row.get(col, "") for col in xlsx_fieldnames])
# Style titres : gras, fond bleu clair, texte blanc
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
# Largeur auto, wrapText et gel de la première ligne/colonne
ws.freeze_panes = ws[2][1]  # freeze first row and first column
for col_idx, col in enumerate(xlsx_fieldnames, 1):
    max_length = len(col)
    for row_cells in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        for cell in row_cells:
            cell.alignment = cell.alignment.copy(wrapText=True)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
# WrapText aussi sur les titres
for cell in ws[1]:
    cell.alignment = cell.alignment.copy(wrapText=True)
# Filtre automatique
ws.auto_filter.ref = ws.dimensions
wb.save("iam_identity_center_report.xlsx")

# Write HTML with DataTables
import pandas as pd

# Remplacer '\n' par '<br>' dans les cellules pour HTML
html_rows = []
for row in rows:
    html_row = {}
    for col in fieldnames:
        val = row[col]
        if isinstance(val, str):
            html_row[col] = val.replace("\n", "<br>")
        else:
            html_row[col] = val
    html_rows.append(html_row)
df = pd.DataFrame(html_rows, columns=fieldnames)
# Only apply string replacement to text columns
text_columns = ["User", "Groups", "AWS Accounts", "Role Details", "Highest Risk Level"]
for col in text_columns:
    if col in df.columns:
        df[col] = df[col].astype(str).str.replace("\n", "<br>", regex=False)
html_table = df.to_html(index=False, escape=False, classes="display nowrap", border=0)
html_template = f"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>AWS IAM Identity Center Report</title>
<link rel="stylesheet" href="https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css">
<style>
table.dataTable thead th {{
    background-color: #4F81BD;
    color: white;
    font-weight: bold;
}}
td, th {{
    white-space: pre-line;
    word-break: break-word;
}}
table {{ width: 100%; }}
</style>
</head>
<body>
<h2>AWS IAM Identity Center Report</h2>
{html_table}
<script src="https://code.jquery.com/jquery-3.7.0.js"></script>
<script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
<script>
$(document).ready(function() {{
    $('table.display').DataTable({{
        scrollX: true,
        paging: true,
        autoWidth: false,
        fixedHeader: true
    }});
}});
</script>
</body>
</html>
"""
with open("iam_identity_center_report.html", "w", encoding="utf-8") as f:
    f.write(html_template)

end_time = datetime.now()
duration = (end_time - start_time).total_seconds()
print("CSV file iam_identity_center_report.csv generated.")
print("XLSX file iam_identity_center_report.xlsx generated.")
print(f"Script started at: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
print(f"Script ended at:   {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
print(f"Total duration:    {duration:.2f} seconds")

# Write JSON output

import json

json_rows = []
for row in rows:
    # Pour JSON, 'Groups' et 'AWS Accounts' sont des listes/objets
    # structurés
    json_row = {
        "User": row["User"],
        "Groups": row.get("_groups_list", []),
        "AWS Accounts": row.get("_aws_accounts_json", []),
    }
    json_rows.append(json_row)
with open("iam_identity_center_report.json", "w", encoding="utf-8") as jf:
    json.dump(json_rows, jf, ensure_ascii=False, indent=2)
print("JSON file iam_identity_center_report.json generated.")
